from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ROOT = Path("/Users/tungyiwu/Desktop/AI/Auto QB/Auto-bot-QB")
DOWNLOADS = Path("/Users/tungyiwu/Downloads")
OUT_DIR = ROOT / "data" / "import-ready"

FILES = {
    "customers": DOWNLOADS / "客戶資料2026031854726.xlsx",
    "vendors": DOWNLOADS / "廠商資料2026031854743.xlsx",
    "products": DOWNLOADS / "商品資料2026031854756.xlsx",
    "sales_returns": DOWNLOADS / "銷退貨彙總表2026031854939.xlsx",
    "profit_analysis": DOWNLOADS / "銷貨利潤分析表2026031855039.xlsx",
}


def clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def load_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = []
    for row in rows[1:]:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        data.append({headers[i]: clean(row[i] if i < len(row) else "") for i in range(len(headers))})
    return data


def write_csv(path: Path, rows: Iterable[dict], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
      writer = csv.DictWriter(f, fieldnames=fieldnames)
      writer.writeheader()
      for row in rows:
          writer.writerow({key: row.get(key, "") for key in fieldnames})


def map_customer_stage(customer_type: str) -> str:
    text = (customer_type or "").strip()
    if "正式" in text:
        return "customer"
    if any(keyword in text for keyword in ["潛在", "詢價", "準客"]):
        return "prospect"
    return "lead"


def build_customer_notes(row: dict) -> str:
    note_parts = []
    pairs = [
        ("電話", row.get("電話")),
        ("傳真", row.get("傳真")),
        ("職稱", row.get("職稱")),
        ("請款客戶", row.get("請款客戶")),
        ("客戶類型", row.get("客戶類型")),
        ("負責人", row.get("負責人")),
        ("客戶等級", row.get("客戶等級")),
        ("首次交易日", row.get("首次交易日")),
        ("合約起始日", row.get("合約起始日")),
        ("合約截止日", row.get("合約截止日")),
    ]
    for label, value in pairs:
        if value:
            note_parts.append(f"{label}:{value}")
    return " | ".join(note_parts)


def convert_customers():
    rows = load_rows(FILES["customers"])
    mapped = []
    for row in rows:
        mapped.append(
            {
                "customer_code": row.get("客戶代號", ""),
                "name": row.get("主聯絡人") or row.get("客戶簡稱", ""),
                "company_name": row.get("客戶簡稱", ""),
                "phone": row.get("手機") or row.get("電話", ""),
                "email": "",
                "tax_id": row.get("統一編號", ""),
                "address": row.get("送貨地址", ""),
                "source": "import",
                "display_name": row.get("客戶簡稱", ""),
                "customer_stage": map_customer_stage(row.get("客戶類型", "")),
                "status": "active",
                "notes": build_customer_notes(row),
            }
        )
    write_csv(
        OUT_DIR / "erp_customers_import.csv",
        mapped,
        [
            "customer_code",
            "name",
            "company_name",
            "phone",
            "email",
            "tax_id",
            "address",
            "source",
            "display_name",
            "customer_stage",
            "status",
            "notes",
        ],
    )


def convert_vendors():
    rows = load_rows(FILES["vendors"])
    mapped = []
    for row in rows:
        mapped.append(
            {
                "vendor_code": row.get("廠商代號", ""),
                "vendor_name": row.get("廠商簡稱", ""),
                "phone": row.get("電話", ""),
                "fax": row.get("傳真", ""),
                "contact_name": row.get("聯絡人", ""),
                "contact_title": row.get("職稱", ""),
                "mobile": row.get("手機", ""),
                "address": row.get("營業地址", ""),
                "tax_id": row.get("統一編號", ""),
            }
        )
    write_csv(
        OUT_DIR / "erp_vendors_import.csv",
        mapped,
        ["vendor_code", "vendor_name", "phone", "fax", "contact_name", "contact_title", "mobile", "address", "tax_id"],
    )


def build_product_description(row: dict) -> str:
    parts = [row.get("品名", ""), row.get("規格一", ""), row.get("規格二", "")]
    return " ".join(str(part).strip() for part in parts if part).strip()


def build_search_text(row: dict) -> str:
    parts = [
        row.get("品號", ""),
        row.get("品名", ""),
        row.get("規格一", ""),
        row.get("規格二", ""),
        row.get("商品分類", ""),
        row.get("主供應商", ""),
    ]
    return " ".join(str(part).strip() for part in parts if part).strip()


def convert_products():
    rows = load_rows(FILES["products"])
    quickbuy_rows = []
    erp_rows = []
    for row in rows:
        description = build_product_description(row)
        category = row.get("商品分類", "") or "other"
        quickbuy_rows.append(
            {
                "item_number": row.get("品號", ""),
                "description": description,
                "tw_retail_price": row.get("零售價", 0) or 0,
                "tw_reseller_price": row.get("優惠價", 0) or 0,
                "product_status": "Current",
                "category": category,
                "replacement_model": "",
                "weight_kg": row.get("單位淨重", 0) or 0,
                "origin_country": "",
                "search_text": build_search_text(row),
            }
        )
        erp_rows.append(
            {
                "item_number": row.get("品號", ""),
                "name": row.get("品名", ""),
                "description": description,
                "brand": row.get("主供應商", "") or "Unknown",
                "unit": row.get("單位", "") or "pcs",
                "cost_price": row.get("標準進價", 0) or 0,
                "list_price": row.get("零售價", 0) or 0,
                "sale_price": row.get("優惠價", 0) or 0,
                "product_status": "active",
                "barcode": row.get("條碼編號", ""),
                "metadata_json": str(
                    {
                        "category": row.get("商品分類", ""),
                        "safety_stock": row.get("安全存量", 0) or 0,
                        "stock_qty": row.get("庫存量", 0) or 0,
                        "supplier": row.get("主供應商", ""),
                    }
                ),
            }
        )

    write_csv(
        OUT_DIR / "quickbuy_products_import.csv",
        quickbuy_rows,
        [
            "item_number",
            "description",
            "tw_retail_price",
            "tw_reseller_price",
            "product_status",
            "category",
            "replacement_model",
            "weight_kg",
            "origin_country",
            "search_text",
        ],
    )
    write_csv(
        OUT_DIR / "erp_products_import.csv",
        erp_rows,
        [
            "item_number",
            "name",
            "description",
            "brand",
            "unit",
            "cost_price",
            "list_price",
            "sale_price",
            "product_status",
            "barcode",
            "metadata_json",
        ],
    )


def convert_sales_returns():
    rows = load_rows(FILES["sales_returns"])
    mapped = []
    for row in rows:
        doc_no = str(row.get("單號", "") or "")
        mapped.append(
            {
                "doc_date": row.get("日期", ""),
                "doc_no": doc_no,
                "doc_type": "return" if doc_no.startswith("退") else "sale",
                "invoice_no": row.get("發票號碼", ""),
                "customer_name": row.get("客戶簡稱", ""),
                "sales_name": row.get("業務姓名", ""),
                "amount": row.get("合計金額", 0) or 0,
                "tax_amount": row.get("稅額", 0) or 0,
                "total_amount": row.get("總金額", 0) or 0,
            }
        )
    write_csv(
        OUT_DIR / "sales_returns_summary_import.csv",
        mapped,
        ["doc_date", "doc_no", "doc_type", "invoice_no", "customer_name", "sales_name", "amount", "tax_amount", "total_amount"],
    )


def convert_profit_analysis():
    rows = load_rows(FILES["profit_analysis"])
    mapped = []
    for row in rows:
        mapped.append(
            {
                "customer_name": row.get("客戶簡稱", ""),
                "doc_date": row.get("日期", ""),
                "doc_no": row.get("單號", ""),
                "sales_name": row.get("業務", ""),
                "amount": row.get("金額", 0) or 0,
                "cost": row.get("成本", 0) or 0,
                "gross_profit": row.get("毛利", 0) or 0,
                "gross_margin": row.get("毛利率", ""),
            }
        )
    write_csv(
        OUT_DIR / "profit_analysis_import.csv",
        mapped,
        ["customer_name", "doc_date", "doc_no", "sales_name", "amount", "cost", "gross_profit", "gross_margin"],
    )


if __name__ == "__main__":
    convert_customers()
    convert_vendors()
    convert_products()
    convert_sales_returns()
    convert_profit_analysis()
    print(f"Converted files written to: {OUT_DIR}")
